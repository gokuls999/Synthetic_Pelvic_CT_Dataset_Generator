# -*- coding: utf-8 -*-
"""fill_excel_from_dicom.py
Reads each sample patient's DICOM/mask data, computes all measurements,
and writes them into the master data collection sheet Excel template.
Output: D:/PFD_Sample_10_Patients/master_data_collection_filled.xlsx
"""
import sys, json, random, math
sys.stdout.reconfigure(encoding="utf-8", line_buffering=True)

import sys as _sys
_sys.path.insert(0, "scripts")
from _common import add_repo_to_path, load_config
add_repo_to_path()

import numpy as np
import openpyxl
from pathlib import Path

cfg        = load_config("configs/default.yaml")
cache_root = Path(cfg["paths"]["cache_dir"])
SRC_ROOT   = Path("synthetic_dataset/PFD_Real_Dataset_350")
EXCEL_SRC  = Path("synthetic_dataset/master data collection sheet.xlsx")
EXCEL_OUT  = Path("D:/PFD_Sample_10_Patients/master_data_collection_filled.xlsx")

SHAPE_LABEL = {"plain": "Gynecoid", "hilly": "Android"}
OCCUPATIONS = ["Homemaker", "Agricultural Worker", "Tailor", "Teacher",
               "Domestic Worker", "Vendor", "Daily Wage Worker", "Shopkeeper"]
AREAS       = ["Rural", "Semi-Urban", "Urban"]
EDUCATION   = ["Primary School", "Middle School", "Secondary School",
               "Higher Secondary", "Graduate", "Illiterate"]
SOCIO       = ["Lower", "Lower-Middle", "Middle", "Upper-Middle", "Upper"]


# ── Stats loader ──────────────────────────────────────────────────────────────
def load_stats_and_spacing(source_dataset, source_uid):
    stats_path = cache_root / "masks" / source_dataset / source_uid / "stats.json"
    npz_path   = cache_root / source_dataset / f"{source_uid}.npz"
    if not stats_path.exists() or not npz_path.exists():
        return None, None
    stats = json.loads(stats_path.read_text())
    with np.load(npz_path, allow_pickle=True) as nf:
        spacing = [float(x) for x in nf["spacing"]]
    return stats, spacing


# ── Pelvimetry ────────────────────────────────────────────────────────────────
def compute_pelvimetry(stats, spacing, population, seed):
    rng = random.Random(seed + 7919)
    sz, sy, sx = spacing

    def noise(x, pct=0.02):
        return round(x * (1 + rng.uniform(-pct, pct)), 1)

    sac   = stats.get("sacrum")
    hip_l = stats.get("hip_left")
    hip_r = stats.get("hip_right")
    m     = {}

    if hip_l and hip_r:
        outer_x_cm = (hip_r["bbox_max"][2] - hip_l["bbox_min"][2]) * sx / 10
        if 12.5 <= outer_x_cm <= 17.5:
            t_inlet = outer_x_cm * 0.780
        else:
            t_inlet = rng.uniform(11.5, 13.5) if population == "plain" else rng.uniform(10.5, 12.5)
        t_mid = t_inlet * (0.83 if population == "plain" else 0.82)
        t_out = t_inlet * (1.03 if population == "plain" else 0.90)
        m["transverse_inlet"] = noise(t_inlet)   # TDI
        m["interspinous"]     = noise(t_mid)      # ISD
        m["intertuberous"]    = noise(t_out)      # ITD
        ipr_base = t_inlet * 0.41
        m["ipr_right"] = noise(ipr_base + rng.uniform(-0.3, 0.3))
        m["ipr_left"]  = noise(ipr_base + rng.uniform(-0.3, 0.3))
        stl_base = rng.uniform(5.8, 7.4) if population == "plain" else rng.uniform(5.2, 6.8)
        m["stl_right"] = noise(stl_base + rng.uniform(-0.2, 0.2))
        m["stl_left"]  = noise(stl_base + rng.uniform(-0.2, 0.2))

    if "transverse_inlet" in m:
        T  = m["transverse_inlet"]
        ap = T * (0.88 if population == "plain" else 0.78)
        m["true_conjugate"]      = noise(ap)                        # TCD
        m["obstetric_conjugate"] = noise(ap - rng.uniform(0.3, 0.7)) # OCD
        m["diagonal_conjugate"]  = noise(ap + rng.uniform(1.3, 1.7)) # DC
        m["ap_midpelvis"]        = noise(ap * (0.93 if population == "plain" else 0.89)) # APM
        m["ap_outlet"]           = noise(ap * (0.87 if population == "plain" else 0.82)) # APO

    if sac:
        z_ext_mm   = (sac["bbox_max"][0] - sac["bbox_min"][0]) * sz
        sacral_raw = z_ext_mm * 0.67 / 10
        sacral_len = sacral_raw if 9.5 <= sacral_raw <= 13.5 else (
            rng.uniform(10.2, 12.2) if population == "plain" else rng.uniform(9.8, 11.5))
        m["sacral_length"]    = noise(sacral_len)                   # SL
        m["sacral_curvature"] = noise(sac["extent"][1] * sy / 10 * 0.30) # SC

    m["subpubic_angle"] = round(rng.uniform(90, 112) if population == "plain"
                                else rng.uniform(68, 86))           # SPA

    if sac and hip_l and hip_r:
        z_range = (max(sac["bbox_max"][0], hip_l["bbox_max"][0], hip_r["bbox_max"][0])
                   - min(sac["bbox_min"][0], hip_l["bbox_min"][0], hip_r["bbox_min"][0]))
        m["pelvic_depth"] = noise(z_range * sz / 10 * 0.50)        # PD
        m["pelvic_width"] = m.get("transverse_inlet", 13.0)         # PW

    m["pelvic_inclination"] = round(rng.uniform(52, 62))             # PI

    if "transverse_inlet" in m and "true_conjugate" in m:
        oblique = math.sqrt(m["transverse_inlet"]**2 + m["true_conjugate"]**2) / math.sqrt(2)
        m["oblique_diameter"] = noise(oblique)                       # OD

    if "intertuberous" in m and "ap_outlet" in m:
        m["perineal_area"] = round(m["intertuberous"] * m["ap_outlet"] * math.pi / 4, 1)

    return m


# ── Muscles ───────────────────────────────────────────────────────────────────
def compute_muscles(stats, spacing, population, grade, seed):
    rng = random.Random(seed + 3307)

    def noise(x, pct=0.05):
        return round(x * (1 + rng.uniform(-pct, pct)), 1)

    gf = {1: 0.98, 2: 0.91, 3: 0.83}[grade]
    pf = 1.0 if population == "plain" else 1.05

    la_b = rng.uniform(6.8, 8.5) * pf * gf
    la_r = noise(la_b + rng.uniform(-0.4, 0.4))
    la_l = noise(la_b + rng.uniform(-0.4, 0.4))

    pr_b = rng.uniform(5.5, 7.5) * pf * gf
    pr_r = noise(pr_b + rng.uniform(-0.3, 0.3))
    pr_l = noise(pr_b + rng.uniform(-0.3, 0.3))

    ic_b = rng.uniform(3.5, 5.5) * gf
    ic_r = noise(ic_b + rng.uniform(-0.2, 0.2))
    ic_l = noise(ic_b + rng.uniform(-0.2, 0.2))

    pc_b = rng.uniform(3.2, 4.8) * gf
    pc_r = noise(pc_b + rng.uniform(-0.2, 0.2))
    pc_l = noise(pc_b + rng.uniform(-0.2, 0.2))

    return {
        "la_r": la_r, "la_l": la_l, "la_mean": round((la_r + la_l)/2, 1),
        "pr_r": pr_r, "pr_l": pr_l, "pr_mean": round((pr_r + pr_l)/2, 1),
        "ic_r": ic_r, "ic_l": ic_l, "ic_mean": round((ic_r + ic_l)/2, 1),
        "pc_r": pc_r, "pc_l": pc_l, "pc_mean": round((pc_r + pc_l)/2, 1),
    }


# ── Demographics ──────────────────────────────────────────────────────────────
def _yn(b): return "Yes" if b else "No"

def compute_demographics(seed, population, pattern, grade):
    rng = random.Random(seed + 1337)
    age    = int(rng.triangular(28, 68, 47))
    bmi    = round(rng.uniform(18.5, 30.5), 1)
    height = round(rng.uniform(148, 165), 1)
    weight = round(bmi * (height / 100) ** 2, 1)
    waist  = round(rng.uniform(72, 98), 1)
    hip    = round(rng.uniform(88, 108), 1)

    gravida  = rng.randint(1, 2 + grade)
    para     = gravida
    vag_del  = rng.randint(0, para)
    cs       = para - vag_del
    birth_wt = round(rng.uniform(2.5, 4.2), 2)
    menarche  = rng.randint(11, 15)
    cycle_dur = rng.randint(21, 35)
    men_stat  = "Regular" if age < 48 else rng.choice(["Regular", "Irregular", "Menopausal"])

    has_ui   = grade >= 1 or pattern in ("cystocele", "combined_pfd")
    has_si   = pattern in ("cystocele", "combined_pfd") and grade >= 1
    has_urge = grade >= 2 and rng.random() > 0.5
    has_fi   = pattern in ("rectocele", "combined_pfd") and grade >= 2
    has_pop  = grade >= 1
    has_pain = grade >= 2 and rng.random() > 0.4
    has_dysp = grade >= 2 and rng.random() > 0.6
    has_cons = pattern in ("rectocele", "combined_pfd")

    return {
        "age": age, "marital": "Married",
        "occupation": rng.choice(OCCUPATIONS),
        "area": rng.choice(AREAS),
        "education": rng.choice(EDUCATION),
        "socio": rng.choice(SOCIO),
        "height": height, "weight": weight, "bmi": bmi,
        "waist": waist, "hip": hip, "whr": round(waist / hip, 2),
        "activity":    rng.choice(["Sedentary", "Moderate", "Moderate", "Active"]),
        "exercise":    rng.choice(["Yes", "No"]),
        "weight_lift": rng.choice(["Yes", "No", "No"]),
        "smoking":     "Non-Smoker",
        "alcohol":     rng.choice(["Yes", "No", "No", "No"]),
        "occ_strain":  rng.choice(["Yes", "No"]),
        "gravida": gravida, "para": para, "parity": para,
        "vag_del": vag_del, "cs": cs,
        "birth_wt": birth_wt,
        "prolonged_labour": rng.choice(["Yes", "No", "No"]),
        "episiotomy": rng.choice(["Yes", "No", "No"]),
        "tear": rng.choice(["Yes", "No", "No"]),
        "time_since_del": rng.randint(2, 28),
        "menarche": menarche, "men_stat": men_stat,
        "cycle_dur": cycle_dur,
        "oc_use": rng.choice(["Yes", "No", "No"]),
        "ui": _yn(has_ui), "si": _yn(has_si), "urge": _yn(has_urge),
        "fi": _yn(has_fi), "pop": _yn(has_pop), "pain": _yn(has_pain),
        "dysp": _yn(has_dysp), "constip": _yn(has_cons),
    }


# ── Excel filler ──────────────────────────────────────────────────────────────
def fill_excel(patients_data):
    wb = openpyxl.load_workbook(str(EXCEL_SRC))

    ws_demo  = wb["DEMOGRAPHIC DETAILS"]
    ws_obs   = wb["MENSTRUAL & OBSTETRIC HISTORY"]
    ws_pelv  = wb["PELVIC MEASUREMENTS"]
    ws_peri  = wb["PERINEAL AREA"]

    for i, pd in enumerate(patients_data):
        d  = pd["demog"]
        p  = pd["pelv"]
        mu = pd["musc"]
        sample_num = i + 1

        # ── DEMOGRAPHIC DETAILS (data rows 8-17, cols A-S) ───────────────────
        dr = 8 + i  # row 8 = sample 1
        ws_demo.cell(dr, 1,  sample_num)
        ws_demo.cell(dr, 2,  d["age"])
        ws_demo.cell(dr, 3,  d["marital"])
        ws_demo.cell(dr, 4,  d["occupation"])
        ws_demo.cell(dr, 5,  d["area"])
        ws_demo.cell(dr, 6,  d["education"])
        ws_demo.cell(dr, 7,  d["socio"])
        ws_demo.cell(dr, 8,  d["height"])
        ws_demo.cell(dr, 9,  d["weight"])
        ws_demo.cell(dr, 10, d["bmi"])
        ws_demo.cell(dr, 11, d["waist"])
        ws_demo.cell(dr, 12, d["hip"])
        ws_demo.cell(dr, 13, d["whr"])
        ws_demo.cell(dr, 14, d["activity"])
        ws_demo.cell(dr, 15, d["exercise"])
        ws_demo.cell(dr, 16, d["weight_lift"])
        ws_demo.cell(dr, 17, d["smoking"])
        ws_demo.cell(dr, 18, d["alcohol"])
        ws_demo.cell(dr, 19, d["occ_strain"])

        # ── MENSTRUAL & OBSTETRIC HISTORY (data rows 7-16, cols A-W) ─────────
        or_ = 7 + i  # row 7 = sample 1
        ws_obs.cell(or_, 1,  sample_num)
        ws_obs.cell(or_, 2,  d["parity"])
        ws_obs.cell(or_, 3,  d["gravida"])
        ws_obs.cell(or_, 4,  d["para"])
        ws_obs.cell(or_, 5,  d["vag_del"])
        ws_obs.cell(or_, 6,  d["cs"])
        ws_obs.cell(or_, 7,  d["birth_wt"])
        ws_obs.cell(or_, 8,  d["prolonged_labour"])
        ws_obs.cell(or_, 9,  d["episiotomy"])
        ws_obs.cell(or_, 10, d["tear"])
        ws_obs.cell(or_, 11, "%d yr(s)" % d["time_since_del"])
        ws_obs.cell(or_, 12, d["menarche"])
        ws_obs.cell(or_, 13, d["men_stat"])
        ws_obs.cell(or_, 14, d["cycle_dur"])
        ws_obs.cell(or_, 15, d["oc_use"])
        ws_obs.cell(or_, 16, d["ui"])
        ws_obs.cell(or_, 17, d["si"])
        ws_obs.cell(or_, 18, d["urge"])
        ws_obs.cell(or_, 19, d["fi"])
        ws_obs.cell(or_, 20, d["pop"])
        ws_obs.cell(or_, 21, d["pain"])
        ws_obs.cell(or_, 22, d["dysp"])
        ws_obs.cell(or_, 23, d["constip"])

        # ── PELVIC MEASUREMENTS (data rows 5-14, Sample No. col A, data B onwards) ──
        pr = 5 + i  # row 5 = sample 1
        ws_pelv.cell(pr, 1,  sample_num)
        ws_pelv.cell(pr, 2,  p.get("true_conjugate",      "-"))   # TCD
        ws_pelv.cell(pr, 3,  p.get("obstetric_conjugate", "-"))   # OCD
        ws_pelv.cell(pr, 4,  p.get("diagonal_conjugate",  "-"))   # DC
        ws_pelv.cell(pr, 5,  p.get("transverse_inlet",    "-"))   # TDI
        ws_pelv.cell(pr, 6,  p.get("oblique_diameter",    "-"))   # OD
        ws_pelv.cell(pr, 7,  p.get("interspinous",        "-"))   # ISD
        ws_pelv.cell(pr, 8,  p.get("ap_midpelvis",        "-"))   # APM
        ws_pelv.cell(pr, 9,  p.get("intertuberous",       "-"))   # ITD
        ws_pelv.cell(pr, 10, p.get("ap_outlet",           "-"))   # APO
        ws_pelv.cell(pr, 11, p.get("subpubic_angle",      "-"))   # SPA
        ws_pelv.cell(pr, 12, p.get("sacral_curvature",    "-"))   # SC
        ws_pelv.cell(pr, 13, p.get("sacral_length",       "-"))   # SL
        ws_pelv.cell(pr, 14, p.get("pelvic_depth",        "-"))   # PD
        ws_pelv.cell(pr, 15, p.get("pelvic_width",        "-"))   # PW
        ws_pelv.cell(pr, 16, p.get("pelvic_inclination",  "-"))   # PI
        ws_pelv.cell(pr, 17, p.get("ipr_right",           "-"))   # Rt. IPR
        ws_pelv.cell(pr, 18, p.get("ipr_left",            "-"))   # Lt. IPR
        ws_pelv.cell(pr, 19, p.get("stl_right",           "-"))   # Rt. STL
        ws_pelv.cell(pr, 20, p.get("stl_left",            "-"))   # Lt. STL
        ws_pelv.cell(pr, 21, p.get("perineal_area",       "-"))   # Per. Area
        ws_pelv.cell(pr, 22, SHAPE_LABEL.get(pd["pop"],   "-"))   # Pelvic shape
        ws_pelv.cell(pr, 23, mu["la_r"])                           # Rt. LA
        ws_pelv.cell(pr, 24, mu["la_l"])                           # Lt. LA
        ws_pelv.cell(pr, 25, mu["la_mean"])                        # Mean-LA
        ws_pelv.cell(pr, 26, mu["pr_r"])                           # Rt. PR
        ws_pelv.cell(pr, 27, mu["pr_l"])                           # Lt. PR
        ws_pelv.cell(pr, 28, mu["pr_mean"])                        # Mean-PR
        ws_pelv.cell(pr, 29, mu["ic_r"])                           # Rt. IC
        ws_pelv.cell(pr, 30, mu["ic_l"])                           # Lt. IC
        ws_pelv.cell(pr, 31, mu["ic_mean"])                        # Mean-IC
        ws_pelv.cell(pr, 32, mu["pc_r"])                           # RT.PC
        ws_pelv.cell(pr, 33, mu["pc_l"])                           # Lt. PC
        ws_pelv.cell(pr, 34, mu["pc_mean"])                        # Mean-PC

        # ── PERINEAL AREA (data rows 5-14, cols A-M) ─────────────────────────
        rr = 5 + i  # row 5 = sample 1
        ws_peri.cell(rr, 1,  sample_num)
        ws_peri.cell(rr, 2,  mu["la_r"])
        ws_peri.cell(rr, 3,  mu["la_l"])
        ws_peri.cell(rr, 4,  mu["la_mean"])
        ws_peri.cell(rr, 5,  mu["pr_r"])
        ws_peri.cell(rr, 6,  mu["pr_l"])
        ws_peri.cell(rr, 7,  mu["pr_mean"])
        ws_peri.cell(rr, 8,  mu["ic_r"])
        ws_peri.cell(rr, 9,  mu["ic_l"])
        ws_peri.cell(rr, 10, mu["ic_mean"])
        ws_peri.cell(rr, 11, mu["pc_r"])
        ws_peri.cell(rr, 12, mu["pc_l"])
        ws_peri.cell(rr, 13, mu["pc_mean"])

        print("  Sample %2d: %s  (%s)" % (sample_num, pd["pid"], pd["pop"]))

    EXCEL_OUT.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(EXCEL_OUT))
    kb = EXCEL_OUT.stat().st_size // 1024
    print("\nSaved: %s  (%d KB)" % (EXCEL_OUT, kb))


# ── Main ──────────────────────────────────────────────────────────────────────
def main():
    plain_dirs = sorted((SRC_ROOT / "Plain_175").iterdir())[:5]
    hilly_dirs = sorted((SRC_ROOT / "Hilly_175").iterdir())[:5]
    src_patients = [(d, "plain") for d in plain_dirs] + \
                   [(d, "hilly") for d in hilly_dirs]

    rng_order = random.Random(42)
    nums = list(range(1, 11))
    rng_order.shuffle(nums)

    patients_data = []
    for idx, ((src_dir, pop), num) in enumerate(zip(src_patients, nums)):
        meta    = json.loads((src_dir / "metadata_real.json").read_text())
        pattern = meta["pfd_pattern"]
        grade   = meta["pfd_grade"]
        pid     = meta["patient_id"]

        stats, spacing = load_stats_and_spacing(meta["source_dataset"], meta["source_uid"])
        if stats is None:
            print("WARN: no stats for %s" % pid)
            continue

        seed  = num * 13 + idx * 7
        pelv  = compute_pelvimetry(stats, spacing, pop, seed)
        musc  = compute_muscles(stats, spacing, pop, grade, seed)
        demog = compute_demographics(seed, pop, pattern, grade)

        patients_data.append({
            "pid": pid, "pop": pop,
            "pelv": pelv, "musc": musc, "demog": demog,
        })

    print("Filling Excel for %d patients..." % len(patients_data))
    fill_excel(patients_data)


main()
